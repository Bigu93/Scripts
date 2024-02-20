from openpyxl import load_workbook
import requests
import logging

# Setup your logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

MAX_ROWS_TO_PROCESS = 100
PRODUCT_ID_COLUMN = "L"  # "product ID"
EAN_COLUMN = "AJ"
MINIATURE_COLUMN = "AK"
PHOTO_COLUMN_1 = "AL"
PHOTO_COLUMN_2 = "AM"
PHOTO_COLUMN_3 = "AN"
LENGTH_OF_INSOLE = "AO"
CODE_COLUMN = "AP"
STARTING_ROW = 439
API_TIMEOUT = 20  # seconds
API_URL = "https://xxx.pl/api/admin/v3/products/products/get"
API_CODE = "https://xxx.pl/api/admin/v3/products/products?productIds="
API_SIZE_URL = "https://xxx.pl/api/admin/v3/products/SKUbyBarcode?productIndices="
API_HEADERS = {
    "Accept": "application/json",
    "X-API-KEY": "xxx",
}


def get_sku(product_id):
    try:
        response = requests.get(
            f"{API_SIZE_URL}{product_id}", headers=API_HEADERS, timeout=API_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logging.error(f"API request error for ID {product_id}: {e}")
        return None

def get_product_displayed_code(product_id):
    try:
        response = requests.get(
            f"{API_CODE}{product_id}", headers=API_HEADERS, timeout=API_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logging.error(f"API request error for ID {product_id}: {e}")
        return None

def get_value_for_size(size, size_chart):
    size_pairs = size_chart.split("/")
    for pair in size_pairs:
        size_value = pair.split("-")
        if len(size_value) == 2 and size_value[0] == str(size):
            return size_value[1]
    return None

def extract_size_names(response):
    size_name = ""
    for result in response.get("results", []):
        for product in result.get("productSkuList", []):
            size_name = product.get("sizeName")
    return size_name

def extract_ean(response):
    ean = ""
    for result in response.get("results", []):
        for product in result.get("productSkuList", []):
            ean = product.get("codeProducer")
    return ean

def extract_code(response):
    code = ""
    for result in response.get("results", []):
        code = result.get("productDisplayedCode")
    return code

def process_size_chart(product_id, size_chart):
    size_response = get_sku(product_id)
    if size_response is not None:
        size_name = extract_size_names(size_response)
    else:
        print("Failed to get response from API")

    size_value = get_value_for_size(size_name, size_chart)
    return size_value

def process_ean(product_id):
    ean_response = get_sku(product_id)
    if ean_response is not None:
        ean = extract_ean(ean_response)
    else:
        print("Failed to get response from API")

    return ean

def process_displayed_code(product_id):
    code_response = get_product_displayed_code(product_id)
    if code_response is not None:
        code = extract_code(code_response)
    else:
        print("Failed to get repsonse from API")

    return code

def get_product_parameters(product_id):
    request_data = {
        "params": {
            "returnElements": ["sizeschart_name", "icon_for_auctions", "pictures"],
            "productIndexes": [{"productIndex": product_id}],
        }
    }
    try:
        response = requests.post(
            API_URL, json=request_data, headers=API_HEADERS, timeout=API_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logging.error(f"API request error for ID {product_id}: {e}")
        return None

def update_parameters_in_sheet(sheet, product_id, row):
    json_response = get_product_parameters(product_id)
    if (
        not json_response
        or "results" not in json_response
        or not json_response["results"]
    ):
        logging.error(f"No results found for ID {product_id}")
        return

    product_info = json_response["results"][0]

    ean_value = process_ean(product_id)
    sheet[f"{EAN_COLUMN}{row}"].value = ean_value

    code_value = process_displayed_code(product_id)
    sheet[f"{CODE_COLUMN}{row}"].value = code_value

    size_chart = product_info.get("sizeChartName", "")
    size_value = process_size_chart(product_id, size_chart)
    sheet[f"{LENGTH_OF_INSOLE}{row}"].value = size_value

    auction_icon_url = product_info.get("productAuctionIcon", {}).get(
        "productAuctionIconLargeUrl", ""
    )
    sheet[f"{MINIATURE_COLUMN}{row}"].value = "https://xxx.pl/" + auction_icon_url

    product_images = product_info.get("productImages", [])
    if len(product_images) >= 3:
        first_image_url = product_images[0].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_1}{row}"].value = first_image_url
        second_image_url = product_images[1].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_2}{row}"].value = second_image_url
        third_image_url = product_images[2].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_3}{row}"].value = third_image_url
    elif len(product_images) == 2:
        first_image_url = product_images[0].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_1}{row}"].value = first_image_url
        second_image_url = product_images[1].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_2}{row}"].value = second_image_url
        third_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_3}{row}"].value = third_image_url
    elif len(product_images) == 1:
        first_image_url = product_images[0].get("productImageLargeUrl", "")
        sheet[f"{PHOTO_COLUMN_1}{row}"].value = first_image_url
        second_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_2}{row}"].value = second_image_url
        third_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_3}{row}"].value = third_image_url
    else:
        first_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_1}{row}"].value = first_image_url
        second_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_2}{row}"].value = second_image_url
        third_image_url = "BRAK ZDJĘCIA"
        sheet[f"{PHOTO_COLUMN_3}{row}"].value = third_image_url


    logging.info(f"Updated parameters for product ID {product_id} in row {row}")

def process_workbook(file_path):
    try:
        wb = load_workbook(file_path, keep_vba=True)
        sheet = wb.active

        processed_rows = 0

        for row in range(STARTING_ROW, sheet.max_row + 1):
            if processed_rows >= MAX_ROWS_TO_PROCESS:
               break
            product_id = sheet[f"{PRODUCT_ID_COLUMN}{row}"].value
            if product_id:
                logging.info(f"Updating parameters for ID {product_id} in row {row}")
                update_parameters_in_sheet(sheet, product_id, row)
                processed_rows += 1

        wb.save(file_path)
        logging.info(f"Workbook processed and saved: {file_path}")
    except Exception as e:
        logging.error(f"Error occurred, saving progress and stopping script: {e}")
        wb.save(file_path)
        logging.info(f"Workbook saved: {file_path}")
        raise  # Reraise the exception to stop the script


file_path = "xxx.xlsm"
process_workbook(file_path)
