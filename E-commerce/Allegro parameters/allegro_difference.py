from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def remove_matching_rows(file1_path, file2_path, start_row=5):
    logging.info("Starting process...")
    wb1 = load_workbook(file1_path)
    sheet1 = wb1.active
    logging.info(f"{file1_path} loaded successfully.")

    wb2 = load_workbook(file2_path, keep_vba=True)
    sheet2 = wb2.active
    logging.info(f"{file2_path} loaded successfully.")

    values_to_check = [sheet1[f"J{row}"].value for row in range(start_row, sheet1.max_row + 1) if sheet1[f"J{row}"].value is not None]
    logging.info("Values collected from the first file.")

    rows_to_delete = []

    for row in range(1, sheet2.max_row + 1):
        cell_value = sheet2[f"L{row}"].value
        if cell_value in values_to_check:
            rows_to_delete.append(row)
    logging.info(f"Found {len(rows_to_delete)} rows to delete.")

    rows_to_delete.sort(reverse=True)

    for row in rows_to_delete:
        sheet2.delete_rows(row)
        logging.info(f"Deleted row: {row}")

    wb2.save(file2_path)
    logging.info(f"Rows removed and {file2_path} saved.")

file1_path = "allegro.xlsm"
file2_path = "allegro2.xlsm"
remove_matching_rows(file1_path, file2_path)