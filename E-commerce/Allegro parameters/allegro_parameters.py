from openpyxl import load_workbook
import requests
import logging

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

MAX_ROWS_TO_PROCESS = 11
PRODUCT_ID_COLUMN = "J"  # "Kod producenta"
PARAMETER_COLUMNS = {
    18: "N",  # "Kolor"
    49: "O",  # "Zapięcie"
    23: "P",  # "Materiał zewnętrzny"
    21: "Q",  # "Wkładka"
    435: "R",  # "Wyściółka"
    6: "S",  # "Rodzaj obcasa"
    1443: "T",  # "Platforma / Podeszwa (cm)"
    3322: "U",  # "Ocieplenie"
    3587: "V",  # "Oryginalne opakowanie producenta"
    3597: "W",  # "Materiał podeszwy"
    3592: "X",  # "Nosek"
    3595: "Y",  # "Wzór dominujący"
    2378: "Z",  # "Wysokość"
    3600: "AA",  # "Cechy dodatkowe"
    3604: "AB",  # "Sezon"
    4772: "AC",  # "Kraj pochodzenia"
    429: "AD",  # "Kategoria"
    2492: "AE",  # "Waga gabarytowa w gramach"
    4286: "AF",  # "Wysokość obcasa/platformy"
}
STARTING_ROW = 2897
API_TIMEOUT = 10  # seconds
API_URL = "https://xxx.pl/api/admin/v3/products/products/get"
API_HEADERS = {
    "Accept": "application/json",
    "X-API-KEY": "XXX",
}


def get_product_parameters(product_id):
    request_data = {
        "params": {
            "returnElements": ["parameters"],
            "productIndexes": [{"productIndex": product_id}],
        }
    }
    try:
        response = requests.post(
            API_URL, json=request_data, headers=API_HEADERS, timeout=API_TIMEOUT
        )
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logging.error(f"API request error for ID {product_id}: {e}")
        return None


def update_parameters_in_sheet(sheet, product_id, row):
    json_response = get_product_parameters(product_id)
    if (
        not json_response
        or "results" not in json_response
        or not json_response["results"]
    ):
        logging.error(f"No results found for ID {product_id}")
        return

    product_parameters = json_response["results"][0].get("productParameters", [])

    expected_param_ids = set(PARAMETER_COLUMNS.keys())

    # Keep track of which parameters we have updated
    updated_param_ids = set()

    for param in product_parameters:
        param_id = param.get("parameterId")
        if param_id in PARAMETER_COLUMNS:
            updated_param_ids.add(param_id)
            # Iterate over parameterValues to find the one with Polish descriptions
            for value in param.get("parameterValues", []):
                polish_value_info = next(
                    (
                        desc
                        for desc in value.get("parameterValueDescriptionsLangData", [])
                        if desc.get("langId") == "pol"
                    ),
                    None,
                )
                if polish_value_info:
                    param_value = polish_value_info.get("parameterValueName")
                    column_letter = PARAMETER_COLUMNS[param_id]
                    sheet[f"{column_letter}{row}"].value = param_value
                    break  # Stop after finding the Polish value

    # If a parameter is not present, leave the cell empty
    missing_param_ids = expected_param_ids - updated_param_ids
    for missing_param_id in missing_param_ids:
        column_letter = PARAMETER_COLUMNS[missing_param_id]
        sheet[f"{column_letter}{row}"].value = "BRAK PARAMETRU"


def process_workbook(file_path):
    try:
        wb = load_workbook(file_path, keep_vba=True)
        sheet = wb.active

        processed_rows = 0

        for row in range(STARTING_ROW, sheet.max_row + 1):
            # if processed_rows >= MAX_ROWS_TO_PROCESS:
            #    break
            product_id = sheet[f"{PRODUCT_ID_COLUMN}{row}"].value
            if product_id:
                logging.info(f"Updating parameters for ID {product_id} in row {row}")
                update_parameters_in_sheet(sheet, product_id, row)
                processed_rows += 1

        wb.save(file_path)
        logging.info(f"Workbook processed and saved: {file_path}")
    except Exception as e:
        logging.error(f"Error occurred, saving progress and stopping script: {e}")
        wb.save(file_path)
        logging.info(f"Workbook saved: {file_path}")
        raise  # Reraise the exception to stop the script


file_path = "allegro.xlsm"
process_workbook(file_path)