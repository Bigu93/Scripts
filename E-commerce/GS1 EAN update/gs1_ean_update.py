from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import requests
import hashlib
import logging
import os


logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

TODAY = datetime.today().strftime("%Y%m%d")
SYSTEM_LOGIN = "LOGIN"
SECRET_KEY = "XXX"

TYPE_PRODUCT_COLUMN = "A"
GTIN_COLUMN = "C"
LANGUAGE_COLUMN = "D"
BRAND_COLUMN = "E"
PRODUCT_NAME_COLUMN = "G"
NETTO_COLUMN = "I"
TYPE_OF_SIZE_COLUMN = "J"
GPC_COLUMN = "O"
COUNTRY_COLUMN = "Q"
STATUS_COLUMN = "R"

TYPE_STR = "Produkt do sprzedaży detalicznej/online (GTIN-13, GTIN-12, GTIN-8)"
LANG_STR = "PL"
BRAND_STR = "XXX"
NETTO_STR = "1"
UNIT_STR = "kg"
GPC_STR = "10001077"
COUNTRY_STR = "Polska"
STATUS_STR = "Aktywny (w sprzedaży)"


STARTING_ROW = 3


def generate_auth_key():
    """Generates an authentication key using the secret key and the current date."""
    combined = TODAY + hashlib.sha1(SECRET_KEY.encode()).hexdigest()
    return hashlib.sha1(combined.encode()).hexdigest()


def get_product_name(ean, user_login, auth_key):
    address = (
        "https://xxx.iai-shop.com/api/?gate=products/getSKUbyBarcode/193/json"
    )
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json;charset=UTF-8",
    }
    request_data = {
        "authenticate": {"userLogin": user_login, "authenticateKey": auth_key},
        "params": {"productIndices": [ean], "searchOnlyInCodeIai": "False"},
    }
    response = requests.post(address, json=request_data, headers=headers)
    response_data = response.json()
    results = response_data.get("results", [])
    if results:
        product_sku_list = results[0].get("productSkuList", [])
        if product_sku_list:
            name = product_sku_list[0].get("productName", "Obuwie")
            size = product_sku_list[0].get("sizeName")

            name = name.strip()
            size = size.strip()
            product_name = f"{name} {size}"

            return product_name
    return "Obuwie"


def add_static_values(sheet, row_num):
    sheet[
        TYPE_PRODUCT_COLUMN + str(row_num)
    ].value = TYPE_STR
    sheet[LANGUAGE_COLUMN + str(row_num)].value = LANG_STR
    sheet[BRAND_COLUMN + str(row_num)].value = BRAND_STR
    sheet[NETTO_COLUMN + str(row_num)].value = NETTO_STR
    sheet[TYPE_OF_SIZE_COLUMN + str(row_num)].value = UNIT_STR
    sheet[GPC_COLUMN + str(row_num)].value = GPC_STR
    sheet[COUNTRY_COLUMN + str(row_num)].value = COUNTRY_STR
    sheet[STATUS_COLUMN + str(row_num)].value = STATUS_STR


def process_workbook(file_path):
    logging.info(f"Starting to process the workbook: {file_path}")

    wb = load_workbook(file_path)
    sheet = wb["MojeGS1"]

    logging.info("Workbook loaded successfully.")

    gtin_col_idx = column_index_from_string(GTIN_COLUMN)
    product_name_col_idx = column_index_from_string(PRODUCT_NAME_COLUMN)

    for row_num in range(STARTING_ROW, sheet.max_row + 1):
        ean_cell = sheet.cell(row=row_num, column=gtin_col_idx)
        product_name_cell = sheet.cell(row=row_num, column=product_name_col_idx)

        add_static_values(sheet, row_num)

        if ean_cell.value is None or ean_cell.value == "":
            logging.info(f"No more EAN codes found, stopping at row {row_num}.")
            break

        ean_str = str(int(ean_cell.value))

        product_name = get_product_name(ean_str, SYSTEM_LOGIN, generate_auth_key())
        logging.info(f"EAN code: {ean_str}, row {row_num}.")

        product_name_cell.value = product_name

    wb.save(file_path)
    logging.info(f"Workbook saved: {file_path}")


def process_all_files_in_folder(folder_path):
    logging.info(f"Scanning for Excel files in folder: {folder_path}")

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(folder_path, filename)
            logging.info(f"Found Excel file: {filename} - Starting to process")

            process_workbook(file_path)

            logging.info(f"Finished processing file: {filename}")
        else:
            logging.info(f"Skipping non-Excel file: {filename}")


folder_path = "cat"
process_all_files_in_folder(folder_path)
