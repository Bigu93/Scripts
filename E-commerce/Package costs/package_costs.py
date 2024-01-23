import argparse
import csv
import logging
import requests
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

PARCEL_NUMBER_COLUMN = "L"
NEW_COLUMNS = ["Netto z panelu", "Brutto z panelu", "VAT"]
NO_PACKAGE_STR = "Nie odnaleziono paczki o podanym numerze"
LACK_OF_DATA_STR = "Brak danych o płatności dla zamówienia"

STARTING_ROW = 2  # for Excel
MAX_ROWS_TO_PROCESS = 1000000
API_TIMEOUT = 10  # seconds


class APIHandler:
    def request_shipping_cost(self, package_nr):
        address = f"https://xxx.pl/api/admin/v3/orders/packages?deliveryPackageNumbers={package_nr}"
        headers = {
            "Accept": "application/json",
            "X-API-KEY": "XXX",
        }
        try:
            response = requests.get(address, headers=headers, timeout=API_TIMEOUT)
            return response
        except Exception as e:
            logging.error(f"Error in API request for package number {package_nr}: {e}")
            return None

    def parse_shipping_cost_response(self, response, package_nr):
        if not response or response.status_code not in [200, 207]:
            status_code = response.status_code if response else "No Response"
            logging.error(f"API request failed with status code {status_code}")
            return f"API request failed with status code {status_code}"

        response_data = response.json()
        results = response_data.get("results", [])

        if results:
            for result in results:
                error_info = result.get("errors", {})
                if error_info.get("faultString") == NO_PACKAGE_STR:
                    log_path = "package_not_found.log"
                    with open(log_path, "a") as log_file:
                        log_file.write(f"{NO_PACKAGE_STR}: {package_nr}\n")
                    return NO_PACKAGE_STR

            first_result = results[0]
            package_dict = first_result.get("deliveryPackage", {})
            shipping_costs = package_dict.get("deliveryPackageParameters", {}).get(
                "shippingCosts", []
            )

            if shipping_costs:
                shipping_cost_net = shipping_costs[0].get("shippingCostNet")
                shipping_cost_gross = shipping_costs[0].get("shippingCostGross")
                shipping_cost_vat = shipping_costs[0].get("shippingCostVat")
                return shipping_cost_net, shipping_cost_gross, shipping_cost_vat
            else:
                return LACK_OF_DATA_STR

        return NO_PACKAGE_STR

    def get_shipping_cost(self, package_nr):
        response = self.request_shipping_cost(package_nr)
        return self.parse_shipping_cost_response(response, package_nr)


class BaseProcessor:
    def __init__(self, api_handler):
        self.api_handler = api_handler

    def format_as_string(self, value):
        """Formats a numeric value as a string with a comma as the decimal separator."""
        if value is not None:
            try:
                value_float = float(value)
                return f"{value_float:.2f}".replace(".", ",")
            except ValueError:
                return None
        return None

    def update_row(self, row, netto, brutto, vat):
        row["Netto z panelu"] = (
            self.format_as_string(netto)
            if netto not in [LACK_OF_DATA_STR, NO_PACKAGE_STR]
            else "-"
        )
        row["Brutto z panelu"] = (
            self.format_as_string(brutto)
            if brutto not in [LACK_OF_DATA_STR, NO_PACKAGE_STR]
            else "-"
        )
        row["VAT"] = (
            self.format_as_string(vat)
            if vat not in [LACK_OF_DATA_STR, NO_PACKAGE_STR]
            else "-"
        )


class ExcelProcessor(BaseProcessor):
    def process_workbook(self, file_path):
        try:
            logging.info(f"Starting to process the workbook: {file_path}")
            wb = load_workbook(file_path)
            sheet = wb.active

            existing_headers = [cell.value for cell in sheet[1]]

            next_available_index = sheet.max_column + 1

            # Dictionary to store the index of new columns
            column_indexes = {
                column: idx
                for idx, column in enumerate(existing_headers, start=1)
                if column in NEW_COLUMNS
            }

            # Append new columns only if they do not already exist
            for column in NEW_COLUMNS:
                if column not in column_indexes:
                    column_indexes[column] = next_available_index
                    sheet.cell(row=1, column=next_available_index).value = column
                    next_available_index += 1

            processed_rows = 0

            for row_num in range(
                STARTING_ROW, min(sheet.max_row + 1, STARTING_ROW + MAX_ROWS_TO_PROCESS)
            ):
                parcel_number = sheet[PARCEL_NUMBER_COLUMN + str(row_num)].value
                if not parcel_number:
                    continue

                logging.info(f"Processing row {row_num}")
                netto, brutto, vat = self.api_handler.get_shipping_cost(parcel_number)

                self.update_row(sheet, row_num, column_indexes, netto, brutto, vat)

                processed_rows += 1

            wb.save(file_path)
            logging.info(f"Workbook saved: {file_path}")

        except Exception as e:
            logging.error(f"Error occurred, saving progress and stopping script: {e}")
            wb.save(file_path)
            logging.info(f"Workbook saved: {file_path}")
            raise  # Reraise the exception to stop the script

    def update_row(self, sheet, row_num, column_indexes, netto, brutto, vat):
        values = [netto, brutto, vat]
        for column, value in zip(NEW_COLUMNS, values):
            cell_value = (
                self.format_as_string(value)
                if value not in [LACK_OF_DATA_STR, NO_PACKAGE_STR]
                else "-"
            )
            sheet.cell(row=row_num, column=column_indexes[column]).value = cell_value


class CSVProcessor(BaseProcessor):
    def process_csv_file(
        self,
        file_path,
        package_nr_field,
        package_nr_transform,
        batch_size=10,
    ):
        try:
            logging.info(f"Starting to process the CSV file: {file_path}")
            all_rows = []
            with open(file_path, mode="r", newline="", encoding="utf-8") as file:
                reader = csv.DictReader(file, delimiter=";")
                fieldnames = reader.fieldnames

                new_columns = ["Netto z panelu", "Brutto z panelu", "VAT"]
                for col in new_columns:
                    if col not in fieldnames:
                        fieldnames.append(col)

                for row in reader:
                    for col in new_columns:
                        if col not in row:
                            row[col] = None
                    all_rows.append(row)

            # Process CSV in batches to handle large files efficiently ~5k or more rows
            for i in range(0, len(all_rows), batch_size):
                # The inner loop processes one batch of rows at a time.
                # 'index' is the row number in the CSV.
                # 'row' is the actual data of the row.
                for index, row in enumerate(all_rows[i : i + batch_size], start=i + 1):
                    package_nr = row.get(package_nr_field)
                    if package_nr:
                        package_nr = package_nr_transform
                        logging.info(
                            f"Processing package number: {package_nr} in row {index}"
                        )
                        result = self.api_handler.get_shipping_cost(package_nr)
                        if result == "Package not found":
                            row["Netto z panelu"] = "-"
                            row["Brutto z panelu"] = "-"
                            row["VAT"] = "-"
                        elif isinstance(result, tuple):
                            netto, brutto, vat = result
                            row["Netto z panelu"] = self.format_as_string(netto)
                            row["Brutto z panelu"] = self.format_as_string(brutto)
                            row["VAT"] = self.format_as_string(vat)
                        else:
                            netto = brutto = vat = None

            with open(file_path, mode="w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
                writer.writeheader()
                writer.writerows(all_rows)

        except Exception as e:
            logging.error(f"Error occurred during CSV file processing: {e}")
            raise


def main():
    parser = argparse.ArgumentParser(
        description="Dodawanie kosztów wysyłki z zamówienia do Excela lub CSV z kosztami przesyłek od kuriera"
    )
    parser.add_argument("-f", "--file", required=True, help="Plik Excel lub CSV")
    parser.add_argument(
        "-c",
        "--courier",
        required=True,
        choices=["DPD", "InPost", "GLS"],
        help="Typ kuriera (DPD, InPost, GLS)",
    )
    args = parser.parse_args()

    file_path = args.file
    courier_type = args.courier
    api_handler = APIHandler()

    if courier_type == "DPD" and file_path.endswith(".xlsx"):
        processor = ExcelProcessor(api_handler)
        processor.process_workbook(file_path)
    elif courier_type in ["InPost", "GLS"] and file_path.endswith(".csv"):
        processor = CSVProcessor(api_handler)
        package_nr_field = "nr" if courier_type == "InPost" else "Nr paczki"
        if courier_type == "InPost":
            package_nr_transform = package_nr_field
        else:
            package_nr_transform = package_nr_field.replace(
                "*", ""
            )  # For GLS, replace '*' with empty string
        processor.process_csv_file(file_path, package_nr_field, package_nr_transform)
    else:
        logging.error(f"Niewspierany kurier: {courier_type} i plik: {file_path}.")
        exit(1)


if __name__ == "__main__":
    main()
