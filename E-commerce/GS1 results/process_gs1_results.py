from openpyxl import load_workbook, Workbook
import logging
import os


logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

CONDITIONS_COLUMN = "X"
CONDITION_1 = "[Nazwa jednoznacznie opisująca produkt] Wyświetlana nazwa produktu występuje wielokrotnie dla różnych numerów GTIN w obrębie Uczestnika"
CONDITION_2 = "Zawartość wiersza wielokrotnie powtarza się w pliku"
EXTRACT_COLUMNS = {"B": None, "F": None}
STARTING_ROW = 2


def process_workbook(file_path, output_data):
    logging.info(f"Starting to process the workbook: {file_path}")

    wb = load_workbook(file_path)
    sheet = wb.active  # First sheet by default

    logging.info("Workbook loaded successfully.")

    for row_num in range(STARTING_ROW, sheet.max_row + 1):
        condition_cell = sheet[CONDITIONS_COLUMN + str(row_num)].value

        if condition_cell in [CONDITION_1, CONDITION_2]:
            extracted_data = [
                sheet[col + str(row_num)].value for col in EXTRACT_COLUMNS
            ]
            output_data.append(extracted_data)

    logging.info(f"Finished processing workbook: {file_path}")


def process_all_files_in_folder(folder_path, output_file):
    output_data = []

    logging.info(f"Scanning for Excel files in folder: {folder_path}")

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(folder_path, filename)
            logging.info(f"Found Excel file: {filename} - Starting to process")

            process_workbook(file_path, output_data)

            logging.info(f"Finished processing file: {filename}")
        else:
            logging.info(f"Skipping non-Excel file: {filename}")

    save_to_excel(output_data, output_file)


def save_to_excel(data, file_name):
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    wb.save(file_name)
    logging.info(f"Output saved to {file_name}")


folder_path = "xxx"  # Folder containing the files to process
output_file = "output.xlsx"  # Name of the output file
process_all_files_in_folder(folder_path, output_file)
